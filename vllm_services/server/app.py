"""
FastAPI classifier server — drop-in replacement for llm_services/classifier.

Same API contract:
- POST /push_segment: accepts {call_uuid, role, text, timestamp}
- WebSocket /ws: broadcasts call_update, transcript, classification_update
- GET /health: server and vLLM health check

LLM + light keyword heuristics on the latest segment for progressive vishing score (can go up or down).
"""

import asyncio
import logging
import json
import math
import random
import httpx
import re
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from concurrent.futures import ThreadPoolExecutor

from fastapi import FastAPI, WebSocket, WebSocketDisconnect, Request, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field
from fastapi.templating import Jinja2Templates

from .. import config
from .. import heuristics
from .. import gold_subprompt_rubric
from ..vllm_client import VLLMClient
from ..actions import ActionManager
from .models import TranscriptSegment, HealthResponse
from . import baseline4

# --- Logging ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger("ClassifierService")

# --- App & State ---
app = FastAPI(title="CyberAspis Classifier", version="1.1.0")

# Robust template path
import os
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATES_DIR = os.path.join(CURRENT_DIR, "templates")
templates = Jinja2Templates(directory=TEMPLATES_DIR)
# Project_v2/data for import_calls (resolve from package dir, fallback to cwd/data)
_data_dir = Path(CURRENT_DIR).resolve().parent.parent / "data"
if not (_data_dir / "vishing_dataset.json").exists():
    _data_dir = Path(os.getcwd()) / "data"
DATA_DIR = _data_dir
VISHING_JSON = DATA_DIR / "vishing_dataset.json"
NON_VISHING_JSON = DATA_DIR / "non_vishing_dataset.json"
NEUTRAL_DIR = DATA_DIR / "neutral"
# Timestamped JSON exports from the dashboard ("Save Results") for fine-tuning / audits
RESULTS_EXPORT_DIR = Path(CURRENT_DIR).resolve().parent.parent / "classifier_run_logs"

# Globals initialized at startup
vllm_client: Optional[VLLMClient] = None
action_manager: Optional[ActionManager] = None
actions_enabled: bool = config.ENABLE_ACTIONS_DEFAULT

MAX_BUFFER_AGE_SECONDS = 30.0


def _import_gt_numeric_for_export(buf: "CallBuffer") -> Any:
    """1 = vishing, 0 = non-vishing, empty string if unknown (XLSX/flat rows)."""
    g = buf.ground_truth_is_vishing
    if g is True:
        return 1
    if g is False:
        return 0
    return ""


def _import_gt_numeric_json(buf: "CallBuffer") -> Optional[int]:
    """JSON: 1 vishing, 0 not, null if no import label."""
    g = buf.ground_truth_is_vishing
    if g is True:
        return 1
    if g is False:
        return 0
    return None


def _parse_export_llm_yes_no(s: str) -> Optional[bool]:
    """
    Best-effort parse of rubric completion to True (YES/ΝΑΙ) / False (NO/ΟΧΙ).
    None if empty, error stub, or unparseable.
    """
    raw = (s or "").strip()
    if not raw:
        return None
    if raw.startswith("[σφάλμα") or raw.startswith("Δεν απέμεινε"):
        return None
    t = raw.translate(str.maketrans("", "", '.,;:!?«»"\'…'))
    for w in t.split():
        wl = w.lower()
        if wl in ("ναι", "ναί", "yes", "y"):
            return True
        if wl in ("οχι", "όχι", "οχί", "no", "n"):
            return False
    if re.search(r"(?iu)\b(yes|ναι)\b", raw):
        return True
    if re.search(r"(?iu)\b(no|οχι|όχι)\b", raw):
        return False
    return None


def _gold_eval_match_import_gt_cells(buf: "CallBuffer", sub_resp: Dict[str, str]) -> Dict[str, str]:
    """
    Per sub-prompt: does LLM YES/NO align with import ground truth (vishing vs non)?
    Same naive rule for all rubric steps: expect YES when GT is vishing, NO when not.
    """
    n = gold_subprompt_rubric.GOLD_SUBPROMPT_COUNT
    gt = buf.ground_truth_is_vishing
    out: Dict[str, str] = {}
    for i in range(1, n + 1):
        raw = (sub_resp.get(f"response_{i}") or "").strip()
        llm_b = _parse_export_llm_yes_no(raw)
        key = f"subprompt_{i}_match"
        if gt is None:
            out[key] = ""
        elif llm_b is None:
            out[key] = "UNPARSED"
        elif llm_b == gt:
            out[key] = "YES"
        else:
            out[key] = "NO"
    return out


def _compute_gold_subprompt_match_performance(
    rows: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """
    From flat gold_eval spreadsheet rows: YES/NO match rate per subprompt vs import GT,
    and pooled overall rate (micro-average over all subprompt×call pairs with a clear YES/NO).
    """
    n = gold_subprompt_rubric.GOLD_SUBPROMPT_COUNT
    per: List[Dict[str, Any]] = []
    tot_yes = tot_no = 0
    for i in range(1, n + 1):
        key = f"subprompt_{i}_match"
        yes = no = unparsed = empty = 0
        for r in rows:
            v = (str(r.get(key) or "")).strip()
            if v == "YES":
                yes += 1
            elif v == "NO":
                no += 1
            elif v == "UNPARSED":
                unparsed += 1
            else:
                empty += 1
        comp = yes + no
        tot_yes += yes
        tot_no += no
        per.append(
            {
                "subprompt_index": i,
                "comparable_yes_no": comp,
                "yes_count": yes,
                "no_count": no,
                "unparsed_count": unparsed,
                "empty_or_no_import_gt": empty,
                "match_rate_percent": (round(100.0 * yes / comp, 2) if comp else None),
            }
        )
    comp_all = tot_yes + tot_no
    return {
        "per_subprompt": per,
        "overall": {
            "comparable_yes_no": comp_all,
            "yes_count": tot_yes,
            "no_count": tot_no,
            "match_rate_percent": (
                round(100.0 * tot_yes / comp_all, 2) if comp_all else None
            ),
        },
        "note": (
            "match_rate_percent = 100 * yes_count / (yes_count + no_count) per subprompt; "
            "empty rows have no import GT or no comparison. "
            "Overall pools all subprompt×call YES+NO cells (micro-average, not mean of subprompt rates)."
        ),
    }


def _fill_gold_match_rates_worksheet(ws: Any, perf: Dict[str, Any], model_name: str) -> None:
    """Second sheet for gold_eval XLSX: table + highlight row for overall."""
    ws.cell(row=1, column=1, value="Gold rubric vs import GT — subprompt match rates")
    ws.cell(row=2, column=1, value="vLLM model")
    ws.cell(row=2, column=2, value=model_name)
    headers = (
        "subprompt",
        "comparable_YES_NO",
        "yes_count",
        "no_count",
        "unparsed_count",
        "empty_no_import_gt",
        "match_rate_%",
    )
    r0 = 4
    for c, h in enumerate(headers, 1):
        ws.cell(row=r0, column=c, value=h)
    r = r0 + 1
    for p in perf.get("per_subprompt") or []:
        ws.cell(row=r, column=1, value=p.get("subprompt_index"))
        ws.cell(row=r, column=2, value=p.get("comparable_yes_no"))
        ws.cell(row=r, column=3, value=p.get("yes_count"))
        ws.cell(row=r, column=4, value=p.get("no_count"))
        ws.cell(row=r, column=5, value=p.get("unparsed_count"))
        ws.cell(row=r, column=6, value=p.get("empty_or_no_import_gt"))
        pct = p.get("match_rate_percent")
        ws.cell(row=r, column=7, value=pct if pct is not None else "")
        r += 1
    o = perf.get("overall") or {}
    ws.cell(row=r, column=1, value="ALL subprompts (pooled)")
    ws.cell(row=r, column=2, value=o.get("comparable_yes_no"))
    ws.cell(row=r, column=3, value=o.get("yes_count"))
    ws.cell(row=r, column=4, value=o.get("no_count"))
    ws.cell(row=r, column=5, value="—")
    ws.cell(row=r, column=6, value="—")
    opct = o.get("match_rate_percent")
    ws.cell(row=r, column=7, value=opct if opct is not None else "")
    ws.cell(row=r + 2, column=1, value=perf.get("note") or "")


def _coerce_ground_truth(label) -> Optional[bool]:
    """
    Dataset label -> is vishing ground truth.
    1 / true = vishing scam; 0 / false = legitimate (non-vishing).
    """
    if label is None:
        return None
    if isinstance(label, bool):
        return label
    try:
        n = int(label)
        if n == 1:
            return True
        if n == 0:
            return False
    except (TypeError, ValueError):
        pass
    if isinstance(label, str):
        s = label.strip().lower()
        if s in ("vishing", "1", "true", "yes", "scam"):
            return True
        if s in ("0", "false", "no", "non_vishing", "clean", "legit", "benign"):
            return False
    return None


def _prediction_positive_for_metrics(cl: Optional[dict]) -> Optional[bool]:
    """
    Same rule as labeled metrics: fraud-positive if score >= FINAL threshold,
    optionally OR risk_status CRITICAL when METRICS_COUNT_CRITICAL_AS_VISHING_POSITIVE.
    Returns None if classification is missing or unusable.
    """
    if not cl or not isinstance(cl, dict):
        return None
    if cl.get("risk_status") == "ERROR":
        return None
    score = cl.get("score")
    if score is None:
        return None
    try:
        pred_pos = baseline4.is_vishing_label(float(score))
    except (TypeError, ValueError):
        return None
    count_crit = getattr(config, "METRICS_COUNT_CRITICAL_AS_VISHING_POSITIVE", True)
    if count_crit and not pred_pos and cl.get("risk_status") == "CRITICAL":
        pred_pos = True
    return pred_pos


def _classification_error_dict(detail: str) -> dict:
    """Unified shape for failed classifications (live + baseline)."""
    return {
        "score": None,
        "vishing_score": None,
        "risk_status": "ERROR",
        "label": "NO",
        "category": "normal",
        "domain": "N/A",
        "analysis": (detail or "classification failed")[:200],
    }


def _enrich_classification_dict(cl: Optional[dict]) -> None:
    """
    Attach pred_positive_fraud_metrics so UI matches export / confusion matrix
    (score ≥ threshold, optionally OR CRITICAL).
    When CRITICAL counts as fraud-positive, align label/category with the VISHING bin for UI/filters.
    """
    if not cl or not isinstance(cl, dict):
        return
    if cl.get("risk_status") == "ERROR":
        cl["pred_positive_fraud_metrics"] = None
        return
    cl["pred_positive_fraud_metrics"] = _prediction_positive_for_metrics(cl)
    count_crit = getattr(config, "METRICS_COUNT_CRITICAL_AS_VISHING_POSITIVE", True)
    if count_crit and cl.get("risk_status") == "CRITICAL":
        cl["label"] = "YES"
        cl["category"] = "vishing"


def _metric_float_for_json(x: Optional[float]) -> Optional[float]:
    if x is None:
        return None
    try:
        xf = float(x)
    except (TypeError, ValueError):
        return None
    if math.isnan(xf) or math.isinf(xf):
        return None
    return round(xf, 4)


def _precision_recall_f1_from_counts(
    tp: int, fp: int, fn: int
) -> Tuple[Optional[float], Optional[float], Optional[float]]:
    precision = (tp / (tp + fp)) if (tp + fp) > 0 else None
    recall = (tp / (tp + fn)) if (tp + fn) > 0 else None
    if precision is None or recall is None:
        f1 = None
    elif precision + recall == 0:
        f1 = None
    else:
        f1 = 2 * precision * recall / (precision + recall)
    return precision, recall, f1


def _roc_auc_binary(y_true: List[int], y_score: List[float]) -> Optional[float]:
    """Wilcoxon / Mann–Whitney AUC; O(n_pos * n_neg), fine for dashboard-sized sets."""
    n = len(y_true)
    if n < 2:
        return None
    pos_i = [i for i in range(n) if y_true[i] == 1]
    neg_i = [i for i in range(n) if y_true[i] == 0]
    if not pos_i or not neg_i:
        return None
    conc = 0.0
    for i in pos_i:
        si = y_score[i]
        for j in neg_i:
            sj = y_score[j]
            if si > sj:
                conc += 1.0
            elif si == sj:
                conc += 0.5
    return conc / (len(pos_i) * len(neg_i))


def compute_confusion_for_call_uids(uids: List[str]) -> dict:
    """
    Binary metrics vs imported dataset labels.
    Predicted positive = score at/above FINAL threshold, optionally OR risk_status CRITICAL.
    Also: precision, recall, F1, ROC-AUC on classifier score vs GT (AUC needs both classes + finite scores).
    """
    tp = tn = fp = fn = 0
    skipped_no_label = 0
    skipped_errors = 0
    y_true: List[int] = []
    y_score: List[float] = []
    for uid in uids:
        buf = CALL_BUFFERS.get(uid)
        if not buf or not buf.latest_classification:
            continue
        gt = buf.ground_truth_is_vishing
        if gt is None:
            skipped_no_label += 1
            continue
        cl = buf.latest_classification
        pred_pos = _prediction_positive_for_metrics(cl)
        if pred_pos is None:
            skipped_errors += 1
            continue
        try:
            s = float(cl.get("score"))
            if math.isnan(s) or math.isinf(s):
                raise ValueError("non-finite score")
        except (TypeError, ValueError):
            skipped_errors += 1
            continue
        y_true.append(1 if gt else 0)
        y_score.append(s)
        if gt:
            if pred_pos:
                tp += 1
            else:
                fn += 1
        else:
            if pred_pos:
                fp += 1
            else:
                tn += 1
    evaluated = tp + tn + fp + fn
    accuracy = (tp + tn) / evaluated if evaluated else None
    precision, recall, f1 = _precision_recall_f1_from_counts(tp, fp, fn)
    precision = _metric_float_for_json(precision)
    recall = _metric_float_for_json(recall)
    f1 = _metric_float_for_json(f1)
    roc_auc = None
    if y_true:
        roc_auc = _metric_float_for_json(_roc_auc_binary(y_true, y_score))
    return {
        "TP": tp,
        "TN": tn,
        "FP": fp,
        "FN": fn,
        "accuracy": round(accuracy, 4) if accuracy is not None else None,
        "precision": precision,
        "recall": recall,
        "f1_score": f1,
        "roc_auc": roc_auc,
        "evaluated": evaluated,
        "skipped_no_ground_truth": skipped_no_label,
        "skipped_errors": skipped_errors,
    }


class CallBuffer:
    def __init__(self, call_uuid: str):
        self.call_uuid = call_uuid
        self.segments: List[dict] = []
        self.last_update = time.time()
        self.last_classification_time = 0.0
        self.is_processing = False
        self.latest_classification: Optional[dict] = None
        self.start_time = time.time()
        self.is_vishing_final = False
        self.triggered_actions: List[dict] = []
        # Set when imported from labeled JSON (vishing_dataset / non_vishing_dataset)
        self.ground_truth_is_vishing: Optional[bool] = None
        # Optional: last behavioral-fraud explainability chain (sub-questions); does not drive classification.
        self.latest_explainability: Optional[dict] = None

    def add_segment(self, segment: dict):
        self.segments.append(segment)
        self.last_update = time.time()

    def get_full_text(self) -> str:
        return " ".join([s["text"] for s in self.segments])

    def latest_segment_text(self) -> str:
        if not self.segments:
            return ""
        return str(self.segments[-1].get("text", "") or "")

    def should_classify(self) -> bool:
        if self.is_processing:
            return False
        if self.is_vishing_final:
            return False

        word_count = len(self.get_full_text().split())
        time_since_last = time.time() - self.last_classification_time
        time_since_update = time.time() - self.last_update

        if (
            word_count >= config.MIN_WORDS_FOR_CLASSIFICATION
            and time_since_last > config.DEBOUNCE_SECONDS
        ):
            return True
        if word_count > 0 and time_since_update > config.DEBOUNCE_SECONDS:
            return True
        return False


CALL_BUFFERS: Dict[str, CallBuffer] = {}
WEBSOCKETS: List[WebSocket] = []


# --- WebSocket Broadcasting ---
async def broadcast(message: dict):
    if not WEBSOCKETS:
        return

    msg_type = message.get("type", "unknown")
    if msg_type != "transcript":  # Avoid spam
        logger.info(f"Broadcasting {msg_type} to {len(WEBSOCKETS)} clients")

    async def send_to_ws(ws):
        try:
            await ws.send_json(message)
            return None
        except Exception as e:
            logger.error(f"Failed to send to WS: {e}")
            return ws

    results = await asyncio.gather(*(send_to_ws(ws) for ws in WEBSOCKETS))
    to_remove = [ws for ws in results if ws is not None]
    for ws in to_remove:
        if ws in WEBSOCKETS:
            WEBSOCKETS.remove(ws)


async def broadcast_call_list():
    logger.info(f"Broadcasting call list. Total buffers: {len(CALL_BUFFERS)}")
    now = time.time()
    max_age = getattr(config, "CALL_LIST_MAX_AGE_SECONDS", 86400.0)
    calls = []
    for c in list(CALL_BUFFERS.values()):
        ref = max(c.last_update, getattr(c, "last_classification_time", None) or 0)
        if now - ref < max_age:
            if c.latest_classification:
                _enrich_classification_dict(c.latest_classification)
            calls.append(
                {
                    "uuid": c.call_uuid,
                    "start_time": c.start_time,
                    "last_update": c.last_update,
                    "classification": c.latest_classification,
                    "is_vishing_final": c.is_vishing_final,
                    "last_classification_time": c.last_classification_time,
                    "is_processing": c.is_processing,
                    "transcripts": c.segments[-500:],  # send up to last 500 segments so full conversation is visible
                    "ground_truth_is_vishing": c.ground_truth_is_vishing,
                    "ground_truth_label": _ground_truth_label(c),
                    "explainability": c.latest_explainability,
                    "triggered_actions": c.triggered_actions,
                }
            )
    await broadcast(
        {"type": "call_update", "calls": calls, "server_time": time.time()}
    )


# --- Classification (LLM-only: domain + vishing score, no heuristics, no accumulator) ---
async def _vllm_complete(prompt: str):
    """
    Call local Mistral-7B-Instruct via vLLM (HTTP API shape similar to OpenAI).

    Uses `/v1/completions` (not chat) to avoid Mistral chat-template issues.
    """
    return await _vllm_complete_with_fallback(prompt)


async def _vllm_complete_with_fallback(prompt: str, max_tokens: Optional[int] = None) -> dict:
    global vllm_client
    if not vllm_client:
        raise RuntimeError("vLLM client not ready")
    mt = (
        max_tokens
        if max_tokens is not None
        else getattr(config, "VLLM_CLASSIFIER_COMPLETION_TOKENS", 96)
    )
    try:
        return await vllm_client.complete(prompt, max_tokens=mt, temperature=0.0)
    except Exception as e1:
        if "404" in str(e1) or "Not Found" in str(e1):
            return await vllm_client.complete_via_chat(
                prompt, max_tokens=mt, temperature=0.0
            )
        raise


def _completion_text_from_result(result: dict) -> str:
    choices = result.get("choices") or []
    first = choices[0] if choices else {}
    completion = first.get("text")
    if completion is None:
        msg = first.get("message") or {}
        completion = msg.get("content") if isinstance(msg, dict) else None
    if completion is None:
        delta = first.get("delta") or {}
        completion = delta.get("content") if isinstance(delta, dict) else None
    return (completion or "").strip()


async def _baseline_classify_single_buffer(
    uid: str,
    buffer: CallBuffer,
    *,
    progress_queue: Optional[asyncio.Queue] = None,
) -> Optional[str]:
    """
    Run baseline mode classification for one buffer; update buffer state.
    Returns None on success, error message string on failure.
    """
    try:
        if buffer.is_vishing_final:
            logger.info(f"Baseline4 skipped {uid}: already finalized as vishing")
            return None
        out = await run_behavior_classification(
            buffer, mode="baseline", progress_queue=progress_queue
        )
        vishing_prob = out["score"]
        analysis = out["analysis"]
        domain = out["domain"]
        risk_status = out["risk_status"]
        prog_meta = out["prog_meta"]
        reviewed_full_conversation = out["reviewed_full_conversation"]
        seg_steps = out["segment_llm_steps"]
        used_prog = out["used_segment_progressive"]
        buffer.latest_classification = {
            "score": round(vishing_prob, 3),
            "prob_yes": round(vishing_prob, 3),
            "vishing_score": round(vishing_prob, 3),
            "risk_status": risk_status,
            "label": "YES" if baseline4.is_vishing_label(vishing_prob) else "NO",
            "category": baseline4.category_from_risk_status(risk_status),
            "domain": domain,
            "analysis": analysis[:200],
            "reviewed_full_conversation": reviewed_full_conversation,
            "heuristic_segment": prog_meta.get("heuristic_segment"),
            "heuristic_full": prog_meta.get("heuristic_full"),
            "segment_llm_steps": seg_steps,
            "used_segment_progressive": used_prog,
        }
        if baseline4.should_stop_reclassification(vishing_prob):
            buffer.is_vishing_final = True
            buffer.latest_classification["finalized"] = True
            buffer.latest_classification["final_reason"] = "vishing_threshold_reached"
        else:
            buffer.latest_classification["finalized"] = False
        _enrich_classification_dict(buffer.latest_classification)
        if action_manager and actions_enabled:
            action = await action_manager.trigger_action(uid, risk_status)
            if action:
                buffer.triggered_actions.append(action)
        buffer.latest_classification["triggered_actions"] = buffer.triggered_actions
        buffer.last_classification_time = time.time()
        buffer.last_update = time.time()
        logger.info(
            f"Baseline4 classified {uid}: domain={domain} prob={vishing_prob:.2f} -> {risk_status} "
            f"(llm_steps={seg_steps}, progressive={used_prog})"
        )
        return None
    except Exception as e:
        logger.error(f"Baseline4 failed for {uid}: {e}", exc_info=True)
        buffer.latest_classification = _classification_error_dict(str(e))
        _enrich_classification_dict(buffer.latest_classification)
        buffer.last_classification_time = time.time()
        buffer.last_update = time.time()
        return str(e)


async def run_behavior_classification(
    buffer: CallBuffer,
    *,
    mode: str,
    progress_queue: Optional[asyncio.Queue] = None,
) -> dict:
    """
    Baseline / live: optionally one LLM call per segment with cumulative transcript
    (so every turn is seen in order), then optional CRITICAL full-context review.
    mode: 'baseline' | 'classify'
    """
    segs = buffer.segments
    full_text = baseline4.segments_full_transcript(segs).strip()
    if not full_text:
        raise ValueError("empty transcript")

    use_progressive = False
    if mode == "baseline":
        use_progressive = config.USE_SEGMENT_PROGRESSIVE_BASELINE
    else:
        use_progressive = (
            config.USE_SEGMENT_PROGRESSIVE_CLASSIFY
            and len(segs) <= config.CLASSIFY_MAX_SEGMENTS_FOR_PROGRESSIVE
        )

    reviewed_full_conversation = False
    critical_triage_applied = False
    segment_steps = 0
    analysis = ""
    segment_count_for_progress = 1  # for CRITICAL-phase progress + UI

    if use_progressive:
        indices = [
            i
            for i in range(len(segs))
            if (str(segs[i].get("text") or "")).strip()
        ]
        if not indices:
            raise ValueError("no non-empty segments")
        segment_count_for_progress = len(indices)
        score_state: Optional[float] = None
        domain = "N/A"
        prog_meta: dict = {}
        for step_i, idx in enumerate(indices):
            cumulative = baseline4.segments_to_transcript(segs, idx).strip()
            if not cumulative:
                continue
            if progress_queue is not None:
                await progress_queue.put(
                    {
                        "segment_step": step_i + 1,
                        "segments_total": len(indices),
                        "phase": "segment_llm",
                    }
                )
            latest_seg = str(segs[idx].get("text") or "")
            prompt = baseline4.build_behavior_prompt(
                cumulative, previous_probability=score_state, use_full_text=False
            )
            result = await _vllm_complete_with_fallback(prompt)
            completion = _completion_text_from_result(result)
            _, step_analysis, llm_prob, domain_llm = baseline4.parse_behavior_completion(
                completion
            )
            segment_steps += 1
            if step_analysis:
                analysis = step_analysis
            d = baseline4.finalize_domain(domain_llm, full_text)
            if d != "N/A":
                domain = d
            score_state, prog_meta = baseline4.finalize_progressive_score(
                score_state, llm_prob, latest_seg, full_text
            )
            if baseline4.should_stop_reclassification(score_state):
                break
        vishing_prob = (
            float(score_state)
            if score_state is not None
            else float(config.STARTING_VISHING_PROBABILITY)
        )
    else:
        # Baseline runs must not blend in stale scores from a prior classification (kills recall).
        previous_score = None
        if mode != "baseline" and buffer.latest_classification:
            previous_score = buffer.latest_classification.get("score")
        text_for_prompt = baseline4.segments_full_transcript(segs)
        prompt = baseline4.build_behavior_prompt(
            text_for_prompt, previous_probability=previous_score, use_full_text=False
        )
        if progress_queue is not None:
            await progress_queue.put(
                {"segment_step": 1, "segments_total": 1, "phase": "single_shot_llm"}
            )
        result = await _vllm_complete_with_fallback(prompt)
        completion = _completion_text_from_result(result)
        _, analysis, vishing_prob, domain_llm = baseline4.parse_behavior_completion(
            completion
        )
        latest_seg = buffer.latest_segment_text()
        vishing_prob, prog_meta = baseline4.finalize_progressive_score(
            previous_score, vishing_prob, latest_seg, full_text
        )
        domain = baseline4.finalize_domain(domain_llm, full_text)
        segment_steps = 1

    risk_status = baseline4.prob_to_risk_status(vishing_prob)
    h_full = heuristics.compute_heuristic_score(full_text)
    if risk_status == "CRITICAL":
        if progress_queue is not None:
            await progress_queue.put(
                {
                    "segment_step": segment_steps,
                    "segments_total": max(segment_count_for_progress, 1),
                    "phase": "critical_review",
                }
            )
        full_tr = baseline4.segments_full_transcript(segs)
        if getattr(config, "USE_CRITICAL_TRIAGE_SECOND_PASS", True):
            # Discrete second opinion: same numeric JSON prompt often leaves score stuck ~0.4–0.5.
            review_prompt = baseline4.build_critical_triage_prompt(full_tr, vishing_prob)
            triage_mt = int(getattr(config, "VLLM_TRIAGE_COMPLETION_TOKENS", 48))
            review_result = await _vllm_complete_with_fallback(
                review_prompt, max_tokens=triage_mt
            )
            review_completion = _completion_text_from_result(review_result)
            resolution, review_analysis = baseline4.parse_critical_triage_completion(
                review_completion
            )
            vishing_prob = baseline4.score_from_critical_triage_resolution(resolution)
            risk_status = baseline4.prob_to_risk_status(vishing_prob)
            if review_analysis:
                analysis = review_analysis
            reviewed_full_conversation = True
            critical_triage_applied = True
        else:
            review_prompt = baseline4.build_behavior_prompt(
                full_tr,
                previous_probability=vishing_prob,
                use_full_text=True,
            )
            review_result = await _vllm_complete_with_fallback(review_prompt)
            review_completion = _completion_text_from_result(review_result)
            _, review_analysis, review_prob, review_domain = baseline4.parse_behavior_completion(
                review_completion
            )
            latest_for_review = buffer.latest_segment_text()
            vishing_prob, prog_meta = baseline4.finalize_progressive_score(
                vishing_prob, review_prob, latest_for_review, full_text
            )
            risk_status = baseline4.prob_to_risk_status(vishing_prob)
            if review_analysis:
                analysis = review_analysis
            rd = baseline4.finalize_domain(review_domain, full_text)
            if rd != "N/A":
                domain = rd
            reviewed_full_conversation = True

    elif (
        mode == "baseline"
        and risk_status == "SAFE"
        and h_full >= config.BASELINE_SAFE_RECHECK_HEURISTIC_MIN
        and vishing_prob < config.BASELINE_SAFE_RECHECK_PROB_BELOW
    ):
        if progress_queue is not None:
            await progress_queue.put(
                {
                    "segment_step": segment_steps,
                    "segments_total": max(segment_count_for_progress, 1),
                    "phase": "safe_recheck_full",
                }
            )
        review_prompt = baseline4.build_behavior_prompt(
            baseline4.segments_full_transcript(segs),
            previous_probability=None,
            use_full_text=True,
        )
        review_result = await _vllm_complete_with_fallback(review_prompt)
        review_completion = _completion_text_from_result(review_result)
        _, review_analysis, review_prob, review_domain = baseline4.parse_behavior_completion(
            review_completion
        )
        vishing_prob = max(vishing_prob, float(review_prob))
        risk_status = baseline4.prob_to_risk_status(vishing_prob)
        if review_analysis:
            analysis = review_analysis
        rd = baseline4.finalize_domain(review_domain, full_text)
        if rd != "N/A":
            domain = rd
        reviewed_full_conversation = True

    elif (
        mode == "classify"
        and getattr(config, "CLASSIFY_SAFE_LONG_RECHECK_ENABLED", True)
        and risk_status == "SAFE"
        and vishing_prob < float(getattr(config, "CLASSIFY_SAFE_RECHECK_PROB_BELOW", 0.22))
        and len(full_text.split()) >= int(getattr(config, "CLASSIFY_SAFE_RECHECK_MIN_WORDS", 220))
        and h_full <= float(getattr(config, "CLASSIFY_SAFE_RECHECK_MAX_HEURISTIC", 0.12))
    ):
        # Long calls: first single-shot pass can miss fraud in omitted middle; full packing recheck.
        if progress_queue is not None:
            await progress_queue.put(
                {
                    "segment_step": segment_steps,
                    "segments_total": max(segment_count_for_progress, 1),
                    "phase": "classify_safe_long_recheck",
                }
            )
        review_prompt = baseline4.build_behavior_prompt(
            baseline4.segments_full_transcript(segs),
            previous_probability=None,
            use_full_text=True,
        )
        review_result = await _vllm_complete_with_fallback(review_prompt)
        review_completion = _completion_text_from_result(review_result)
        _, review_analysis, review_prob, review_domain = baseline4.parse_behavior_completion(
            review_completion
        )
        vishing_prob = max(vishing_prob, float(review_prob))
        risk_status = baseline4.prob_to_risk_status(vishing_prob)
        if review_analysis:
            analysis = review_analysis
        rd = baseline4.finalize_domain(review_domain, full_text)
        if rd != "N/A":
            domain = rd
        reviewed_full_conversation = True

    # High keyword heuristic + low LLM score → floor (includes scores in SAFE band; fixes FN).
    if (
        h_full >= config.RECALL_HEURISTIC_FLOOR_THRESHOLD
        and vishing_prob <= config.RECALL_HEURISTIC_FLOOR_MAX_LLM
    ):
        vishing_prob = max(vishing_prob, config.RECALL_HEURISTIC_FLOOR_VALUE)
    risk_status = baseline4.prob_to_risk_status(vishing_prob)

    # Many missed vishing calls: SAFE + low score + Banking/Utilities/Education/Account Verification
    # (benign ~0.2 scores are usually Neutral/N/A — triage second opinion).
    if (
        getattr(config, "USE_SAFE_DOMAIN_TRIAGE_RECHECK", True)
        and not critical_triage_applied
        and risk_status == "SAFE"
        and vishing_prob < float(getattr(config, "SAFE_DOMAIN_TRIAGE_MAX_PROB", 0.42))
        and domain in getattr(config, "SAFE_DOMAIN_TRIAGE_DOMAINS", frozenset())
    ):
        if progress_queue is not None:
            await progress_queue.put(
                {
                    "segment_step": segment_steps,
                    "segments_total": max(segment_count_for_progress, 1),
                    "phase": "safe_domain_triage",
                }
            )
        full_tr = baseline4.segments_full_transcript(segs)
        review_prompt = baseline4.build_critical_triage_prompt(full_tr, vishing_prob)
        triage_mt = int(getattr(config, "VLLM_TRIAGE_COMPLETION_TOKENS", 48))
        review_result = await _vllm_complete_with_fallback(
            review_prompt, max_tokens=triage_mt
        )
        review_completion = _completion_text_from_result(review_result)
        resolution, review_analysis = baseline4.parse_critical_triage_completion(
            review_completion
        )
        vishing_prob = baseline4.score_from_critical_triage_resolution(resolution)
        risk_status = baseline4.prob_to_risk_status(vishing_prob)
        if review_analysis:
            analysis = review_analysis
        reviewed_full_conversation = True
        critical_triage_applied = True

    domain = baseline4.finalize_domain(domain, full_text)

    vishing_prob = baseline4.apply_low_signal_short_call_cap(
        vishing_prob, full_text, h_full
    )
    risk_status = baseline4.prob_to_risk_status(vishing_prob)

    return {
        "score": vishing_prob,
        "analysis": analysis or "",
        "domain": domain,
        "risk_status": risk_status,
        "prog_meta": prog_meta,
        "reviewed_full_conversation": reviewed_full_conversation,
        "segment_llm_steps": segment_steps,
        "used_segment_progressive": use_progressive,
    }


async def _report_quality_event(call_uuid: str, classification: dict):
    """Fire-and-forget: POST classification event to quality_service if configured."""
    quality_url = getattr(config, "QUALITY_SERVICE_URL", "")
    if not quality_url:
        return
    try:
        async with httpx.AsyncClient(timeout=2.0) as client:
            await client.post(
                f"{quality_url}/calls/{call_uuid}/events",
                json={
                    "source": "vllm_services",
                    "event_type": "classification_result",
                    "payload": classification,
                },
            )
    except Exception:
        pass


async def classify_call(buffer: CallBuffer):
    global vllm_client

    buffer.is_processing = True
    await broadcast(
        {
            "type": "classification_update",
            "call_uuid": buffer.call_uuid,
            "classification": buffer.latest_classification,
            "last_classification_time": buffer.last_classification_time,
            "is_processing": True,
            "server_time": time.time(),
        }
    )

    try:
        if not vllm_client:
            logger.error("vLLM client not ready")
            buffer.latest_classification = _classification_error_dict("vLLM client not ready")
            _enrich_classification_dict(buffer.latest_classification)
            buffer.last_classification_time = time.time()
            buffer.last_update = time.time()
            return
        text = buffer.get_full_text().strip()
        if not text:
            logger.warning(f"Empty text for {buffer.call_uuid}")
            buffer.latest_classification = _classification_error_dict("empty transcript")
            _enrich_classification_dict(buffer.latest_classification)
            buffer.last_classification_time = time.time()
            buffer.last_update = time.time()
            return
        out = await run_behavior_classification(buffer, mode="classify")
        vishing_prob = out["score"]
        analysis = out["analysis"]
        domain = out["domain"]
        risk_status = out["risk_status"]
        prog_meta = out["prog_meta"]
        reviewed_full_conversation = out["reviewed_full_conversation"]
        segment_steps = out["segment_llm_steps"]
        used_prog = out["used_segment_progressive"]

        classification = {
            "score": round(vishing_prob, 3),
            "prob_yes": round(vishing_prob, 3),
            "vishing_score": round(vishing_prob, 3),
            "risk_status": risk_status,
            "label": "YES" if baseline4.is_vishing_label(vishing_prob) else "NO",
            "category": baseline4.category_from_risk_status(risk_status),
            "domain": domain,
            "analysis": analysis[:200],
            "reviewed_full_conversation": reviewed_full_conversation,
            "heuristic_segment": prog_meta.get("heuristic_segment"),
            "heuristic_full": prog_meta.get("heuristic_full"),
            "segment_llm_steps": segment_steps,
            "used_segment_progressive": used_prog,
        }
        if baseline4.should_stop_reclassification(vishing_prob):
            buffer.is_vishing_final = True
            classification["finalized"] = True
            classification["final_reason"] = "vishing_threshold_reached"
        else:
            classification["finalized"] = False
        _enrich_classification_dict(classification)
        if action_manager and actions_enabled:
            action = await action_manager.trigger_action(buffer.call_uuid, risk_status)
            if action:
                buffer.triggered_actions.append(action)
        classification["triggered_actions"] = buffer.triggered_actions
        buffer.latest_classification = classification
        buffer.last_classification_time = time.time()
        buffer.last_update = time.time()

        asyncio.create_task(_report_quality_event(buffer.call_uuid, classification))

        await broadcast(
            {
                "type": "classification_update",
                "call_uuid": buffer.call_uuid,
                "classification": classification,
                "last_classification_time": buffer.last_classification_time,
                "triggered_actions": buffer.triggered_actions,
                "is_processing": False,
                "server_time": time.time(),
            }
        )
        logger.info(f"Classification for {buffer.call_uuid}: domain={domain} score={vishing_prob:.2f} -> {risk_status}")

    except Exception as e:
        logger.error(f"Classification Error: {e}", exc_info=True)
        buffer.latest_classification = _classification_error_dict(str(e))
        _enrich_classification_dict(buffer.latest_classification)
        buffer.last_classification_time = time.time()
        buffer.last_update = time.time()
    finally:
        buffer.is_processing = False
        await broadcast(
            {
                "type": "classification_update",
                "call_uuid": buffer.call_uuid,
                "classification": buffer.latest_classification,
                "last_classification_time": buffer.last_classification_time,
                "is_processing": False,
                "server_time": time.time(),
            }
        )


# --- Endpoints ---
@app.get("/")
async def get_index(request: Request):
    return templates.TemplateResponse(request, "index.html")


@app.get("/calls/{call_uuid}")
async def get_call_status(call_uuid: str):
    buf = CALL_BUFFERS.get(call_uuid)
    if buf is None:
        raise HTTPException(status_code=404, detail="call not found")
    return {
        "call_uuid": call_uuid,
        "latest_classification": buf.latest_classification,
        "is_processing": buf.is_processing,
        "n_segments": len(buf.segments),
    }


@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await websocket.accept()
    logger.info(f"New WebSocket connection from {websocket.client}")
    WEBSOCKETS.append(websocket)
    try:
        await broadcast_call_list()
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        logger.info(f"WebSocket disconnected: {websocket.client}")
        if websocket in WEBSOCKETS:
            WEBSOCKETS.remove(websocket)
    except Exception as e:
        logger.error(f"WebSocket Error: {e}")
        if websocket in WEBSOCKETS:
            WEBSOCKETS.remove(websocket)


@app.post("/push_segment")
async def push_segment(segment: TranscriptSegment):
    logger.info(f"Received segment for {segment.call_uuid}: {segment.text[:50]}...")
    is_new_call = segment.call_uuid not in CALL_BUFFERS

    if is_new_call:
        CALL_BUFFERS[segment.call_uuid] = CallBuffer(segment.call_uuid)
        logger.info(f"New call detected: {segment.call_uuid}")

    buffer = CALL_BUFFERS[segment.call_uuid]
    seg_dict = segment.model_dump()
    buffer.add_segment(seg_dict)

    if is_new_call:
        await broadcast_call_list()

    await broadcast(
        {
            "type": "transcript",
            "call_uuid": segment.call_uuid,
            "segment": seg_dict,
            "server_time": time.time(),
        }
    )

    if buffer.should_classify():
        logger.info(f"Triggering classification for {segment.call_uuid}")
        asyncio.create_task(classify_call(buffer))

    return {"status": "ok"}


def _rows_from_labeled_json(path: Path, label_name: str) -> List[dict]:
    """Load call_id + text + label rows from vishing_dataset.json / non_vishing_dataset.json."""
    bucket: List[dict] = []
    if not path.exists():
        logger.warning(f"Dataset not found: {path}")
        return bucket
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        return bucket
    default_lbl = 1 if label_name == "vishing" else 0
    for item in data:
        if not isinstance(item, dict):
            continue
        call_id = item.get("call_id")
        text = item.get("text", "")
        if not call_id or not str(text).strip():
            continue
        bucket.append({
            "call_id": str(call_id),
            "text": str(text),
            "label": item.get("label", default_lbl),
        })
    return bucket


def _neutral_json_file_to_row(path: Path) -> Optional[dict]:
    """One conversation from data/neutral/*.json (dialogue[] with text lines)."""
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError):
        return None
    dialogue = data.get("dialogue") if isinstance(data, dict) else None
    if not isinstance(dialogue, list):
        return None
    lines: List[str] = []
    for turn in dialogue:
        if not isinstance(turn, dict):
            continue
        t = turn.get("text")
        if t is None:
            continue
        s = str(t).strip()
        if s:
            lines.append(s)
    if not lines:
        return None
    call_id = path.stem
    return {"call_id": call_id, "text": "\n".join(lines), "label": None}


def _rows_from_neutral_dir(neutral_dir: Path) -> List[dict]:
    rows: List[dict] = []
    if not neutral_dir.is_dir():
        return rows
    for path in sorted(neutral_dir.glob("*.json")):
        row = _neutral_json_file_to_row(path)
        if row:
            rows.append(row)
    return rows


def _round_robin_take(pools: List[List[dict]], limit_total: int) -> Tuple[List[dict], List[int]]:
    """
    Take up to limit_total rows, rotating across non-empty pools (fair mix).
    Returns (rows, per_pool_taken_counts aligned with pools input).
    """
    n = len(pools)
    counts = [0] * n
    out: List[dict] = []
    if not pools or limit_total <= 0:
        return out, counts
    idx = [0] * n
    pi = 0
    while len(out) < limit_total:
        progressed = False
        for _ in range(n):
            p = pi % n
            pi += 1
            if idx[p] < len(pools[p]):
                out.append(pools[p][idx[p]])
                idx[p] += 1
                counts[p] += 1
                progressed = True
                if len(out) >= limit_total:
                    break
        if not progressed:
            break
    return out, counts


def _load_conversations(
    limit_total: int,
    *,
    include_vishing: bool = True,
    include_non_vishing: bool = True,
    include_neutral: bool = False,
) -> Tuple[List[dict], dict]:
    """
    Load up to limit_total conversations from selected dataset pools.
    Shuffles each pool; mixes selected pools in round-robin order.
    """
    vishing_rows = _rows_from_labeled_json(VISHING_JSON, "vishing")
    non_rows = _rows_from_labeled_json(NON_VISHING_JSON, "non_vishing")
    neutral_rows = _rows_from_neutral_dir(NEUTRAL_DIR)
    random.shuffle(vishing_rows)
    random.shuffle(non_rows)
    random.shuffle(neutral_rows)

    pool_defs: List[Tuple[str, List[dict]]] = []
    if include_vishing:
        pool_defs.append(("vishing", vishing_rows))
    if include_non_vishing:
        pool_defs.append(("non_vishing", non_rows))
    if include_neutral:
        pool_defs.append(("neutral", neutral_rows))

    pools = [p[1] for p in pool_defs]
    names = [p[0] for p in pool_defs]
    out, take_counts = _round_robin_take(pools, limit_total)

    stats: Dict[str, Any] = {
        "requested": limit_total,
        "imported": len(out),
        "vishing_pool_size": len(vishing_rows),
        "non_vishing_pool_size": len(non_rows),
        "neutral_pool_size": len(neutral_rows),
        "from_vishing_file": 0,
        "from_non_vishing_file": 0,
        "from_neutral_dir": 0,
        "pools_selected": names,
    }
    for name, c in zip(names, take_counts):
        if name == "vishing":
            stats["from_vishing_file"] = c
        elif name == "non_vishing":
            stats["from_non_vishing_file"] = c
        elif name == "neutral":
            stats["from_neutral_dir"] = c
    return out, stats


def _text_to_segments(text: str) -> List[dict]:
    raw = [s.strip() for s in text.split("\n") if s.strip()]
    segments = []
    for i, seg in enumerate(raw):
        role = "agent" if i % 2 == 0 else "caller"
        segments.append({"role": role, "text": seg, "timestamp": time.time()})
    return segments


def _parse_bool_flag(val: Any, default: bool) -> bool:
    if val is None:
        return default
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        return val.strip().lower() in ("1", "true", "yes", "on")
    return default


@app.post("/import_calls")
async def import_calls(request: Request):
    """Import N conversations from dataset into CALL_BUFFERS.

    Body: {"limit": N, "include_vishing": bool, "include_non_vishing": bool, "include_neutral": bool}.
    Defaults: vishing + non_vishing on, neutral off (legacy behaviour).
    """
    try:
        body = await request.json()
    except Exception:
        body = {}
    if not isinstance(body, dict):
        body = {}
    try:
        limit = int(body.get("limit", 10))
    except (TypeError, ValueError):
        limit = 10
    limit = max(1, min(limit, 500))
    include_vishing = _parse_bool_flag(body.get("include_vishing"), True)
    include_non_vishing = _parse_bool_flag(body.get("include_non_vishing"), True)
    include_neutral = _parse_bool_flag(body.get("include_neutral"), False)
    if not (include_vishing or include_non_vishing or include_neutral):
        return {
            "ok": False,
            "error": "Επιλέξτε τουλάχιστον μία κατηγορία (vishing, μη vishing ή ουδέτερες).",
            "imported": 0,
        }
    try:
        conversations, load_stats = _load_conversations(
            limit,
            include_vishing=include_vishing,
            include_non_vishing=include_non_vishing,
            include_neutral=include_neutral,
        )
    except Exception as e:
        logger.exception("Load conversations failed")
        return {"ok": False, "error": f"Load failed: {e}", "imported": 0}
    if not conversations:
        return {
            "ok": False,
            "error": (
                "Δεν βρέθηκαν δεδομένα για τις επιλεγμένες κατηγορίες "
                "(έλεγχος: data/vishing_dataset.json, data/non_vishing_dataset.json, data/neutral/*.json)."
            ),
            "imported": 0,
        }
    try:
        for conv in conversations:
            call_id = conv["call_id"]
            if call_id not in CALL_BUFFERS:
                CALL_BUFFERS[call_id] = CallBuffer(call_id)
            buffer = CALL_BUFFERS[call_id]
            buffer.ground_truth_is_vishing = _coerce_ground_truth(conv.get("label"))
            for seg in _text_to_segments(conv["text"]):
                buffer.add_segment(seg)
    except Exception as e:
        logger.exception("Import failed")
        return {"ok": False, "error": str(e), "imported": 0}
    await broadcast_call_list()
    logger.info(
        "Imported %s conversations (v=%s, non_v=%s, neutral=%s; pool sizes v=%s n=%s neu=%s)",
        len(conversations),
        load_stats.get("from_vishing_file"),
        load_stats.get("from_non_vishing_file"),
        load_stats.get("from_neutral_dir"),
        load_stats.get("vishing_pool_size"),
        load_stats.get("non_vishing_pool_size"),
        load_stats.get("neutral_pool_size"),
    )
    return {"ok": True, "imported": len(conversations), "load_stats": load_stats}


@app.post("/delete_calls")
async def delete_all_calls():
    """Remove all calls from the dashboard."""
    count = len(CALL_BUFFERS)
    CALL_BUFFERS.clear()
    await broadcast_call_list()
    logger.info(f"Deleted all calls ({count})")
    return {"ok": True, "deleted": count}


_EXPORT_FILENAME_SAFE = re.compile(
    r"^classifier_run_\d{8}T\d{6}Z(_gold_eval)?\.(json|xlsx)$"
)


class ExplainabilityStreamBody(BaseModel):
    limit: int = Field(500, ge=1, le=500)
    custom_prompt: Optional[str] = Field(
        None,
        max_length=120000,
        description="If non-empty (and custom_subprompts is absent), one LLM call per call with this + transcript only (no classifier/heuristics).",
    )
    custom_subprompts: Optional[List[str]] = Field(
        None,
        description="1..CUSTOM_EXPLAIN_MAX_SUBPROMPTS non-empty strings — one Mistral explainability turn each (no classifier).",
    )
    custom_prompt_preserve_output: bool = Field(
        False,
        description="With custom_prompt only: skip prose sanitizers (use for JSON outputs).",
    )
    gold_eval_run: bool = Field(
        False,
        description="If true: run fixed rubric sub-prompts only (gold eval); marks export for GT columns.",
    )


def _gold_eval_explain_fields(gold_eval_run: bool) -> Dict[str, Any]:
    if not gold_eval_run:
        return {}
    return {
        "gold_eval": True,
        "gold_rubric_id": gold_subprompt_rubric.RUBRIC_ID,
        "gold_subprompt_keys": list(gold_subprompt_rubric.SUBPROMPT_KEYS),
    }


def _normalize_explainability_custom_subprompts(raw: Optional[List[str]]) -> Optional[List[str]]:
    if raw is None:
        return None
    if not isinstance(raw, list):
        return None
    mx = max(1, int(getattr(config, "CUSTOM_EXPLAIN_MAX_SUBPROMPTS", 12) or 12))
    if len(raw) < 1 or len(raw) > mx:
        return None
    out: List[str] = []
    for x in raw:
        s = (str(x) if x is not None else "").strip()
        if not s:
            return None
        out.append(s[:120000])
    return out


_ILLEGAL_OPENXML_TEXT_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _sanitize_openpyxl_cell_str(s: str) -> str:
    """Strip control chars that make openpyxl / Excel reject or corrupt cells."""
    if not s:
        return ""
    return _ILLEGAL_OPENXML_TEXT_CHARS.sub("", s)


def _cell_value_for_xlsx(val: Any) -> Any:
    """openpyxl-safe scalar for a worksheet cell."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val
    return _sanitize_openpyxl_cell_str(str(val))


_XLSX_CELL_MAX_LEN = 31000


def _xlsx_custom_sub_column_names() -> Tuple[str, ...]:
    mx = max(1, int(getattr(config, "CUSTOM_EXPLAIN_MAX_SUBPROMPTS", 12) or 12))
    cols: List[str] = []
    for i in range(1, mx + 1):
        cols.append(f"subprompt_{i}")
        cols.append(f"response_{i}")
    return tuple(cols)


def _xlsx_header_display_name(col_key: str) -> str:
    """Readable first-row header for custom sub-prompt columns."""
    m = re.match(r"^subprompt_(\d+)$", col_key)
    if m:
        return f"Sub-prompt {m.group(1)}"
    m = re.match(r"^response_(\d+)$", col_key)
    if m:
        return f"Response {m.group(1)}"
    if col_key == "import_gt_0_1":
        return "Import GT (1=vishing, 0=not)"
    m = re.match(r"^subprompt_(\d+)_match$", col_key)
    if m:
        return f"Subprompt {m.group(1)} match vs import GT (YES/NO)"
    m = re.match(r"^gt_(\d+)$", col_key)
    if m:
        return f"GT ΝΑΙ/ΟΧΙ {m.group(1)} (χειροκίνητα)"
    if col_key == "annotator_notes":
        return "Σημειώσεις annotator"
    return col_key


def _blank_subprompt_response_cells() -> Dict[str, str]:
    return {k: "" for k in _xlsx_custom_sub_column_names()}


def _xlsx_gold_gt_column_names() -> Tuple[str, ...]:
    n = gold_subprompt_rubric.GOLD_SUBPROMPT_COUNT
    return tuple(f"gt_{i}" for i in range(1, n + 1)) + ("annotator_notes",)


def _xlsx_gold_rubric_mid_columns() -> Tuple[str, ...]:
    """
    gold_eval sheet: subprompt_i, response_i, subprompt_i_match (YES/NO vs import GT) for each rubric step,
    then any remaining subprompt/response slots up to CUSTOM_EXPLAIN_MAX_SUBPROMPTS (no match columns).
    """
    n = gold_subprompt_rubric.GOLD_SUBPROMPT_COUNT
    mx = max(1, int(getattr(config, "CUSTOM_EXPLAIN_MAX_SUBPROMPTS", 12) or 12))
    cols: List[str] = []
    for i in range(1, n + 1):
        cols.extend([f"subprompt_{i}", f"response_{i}", f"subprompt_{i}_match"])
    for i in range(n + 1, mx + 1):
        cols.extend([f"subprompt_{i}", f"response_{i}"])
    return tuple(cols)


def _xlsx_export_columns_core_gold_eval() -> Tuple[str, ...]:
    """Core columns plus explicit numeric import GT for gold_eval XLSX."""
    lst = list(_XLSX_EXPORT_COLUMNS_CORE)
    ix = lst.index("ground_truth") + 1
    lst.insert(ix, "import_gt_0_1")
    return tuple(lst)


def _blank_gt_cells() -> Dict[str, str]:
    return {k: "" for k in _xlsx_gold_gt_column_names()}


def _gold_gt_json_fields(export_variant: str) -> Dict[str, Any]:
    if export_variant != "gold_eval":
        return {}
    return dict(_blank_gt_cells())


def _subprompt_response_cells_for_xlsx(ex: Optional[dict]) -> Dict[str, str]:
    """
    One column pair per chained custom explain step: subprompt_N = user text, response_N = llm_raw.
    """
    out: Dict[str, str] = _blank_subprompt_response_cells()
    if not ex:
        return out
    mx = max(1, int(getattr(config, "CUSTOM_EXPLAIN_MAX_SUBPROMPTS", 12) or 12))
    steps = ex.get("steps") or []
    triples: List[Tuple[int, str, str]] = []
    for s in steps:
        sk = str(s.get("step_key") or "")
        if not sk.startswith("custom_sub_"):
            continue
        try:
            idx = int(sk[len("custom_sub_") :])
        except ValueError:
            continue
        q = str(s.get("llm_question") or "").strip()
        r = str(s.get("llm_raw") or "").strip()
        triples.append((idx, q, r))
    triples.sort(key=lambda t: t[0])
    for display_i, (_, q, r) in enumerate(triples, start=1):
        if display_i > mx:
            break
        out[f"subprompt_{display_i}"] = _sanitize_openpyxl_cell_str(q[:_XLSX_CELL_MAX_LEN])
        out[f"response_{display_i}"] = _sanitize_openpyxl_cell_str(r[:_XLSX_CELL_MAX_LEN])
    return out


def _explain_step_llm_question_baseline(spec: Dict[str, Any]) -> str:
    """Substantive instruction the LLM saw for this explain step (for UI/export)."""
    f = str(spec.get("focus") or "").strip()
    if f:
        return f
    return str(spec.get("title_el") or "").strip()


def _explainability_export_blob(buf: CallBuffer) -> Optional[Dict[str, Any]]:
    ex = buf.latest_explainability
    if not ex:
        return None
    steps = ex.get("steps") or []
    parts: List[str] = []
    for s in steps:
        title = str(s.get("step_title") or s.get("step_key") or "")
        q = str(s.get("llm_question") or "").strip()
        raw = str(s.get("llm_raw") or "")
        chunk = f"[{title}]"
        if q:
            chunk += f"\nLLM Question:\n{q}"
        chunk += f"\nLLM Response:\n{raw}"
        parts.append(chunk)
    return {
        "mode": ex.get("mode", "baseline_steps"),
        "llm_only": bool(ex.get("llm_only")),
        "custom_prompt_preview": ex.get("custom_prompt_preview"),
        "custom_subprompts_preview": ex.get("custom_subprompts_preview"),
        "updated_at": ex.get("updated_at"),
        "steps_count": len(steps),
        "combined_llm_text": "\n\n---\n\n".join(parts),
        "steps": [dict(s) for s in steps],
    }


def _build_export_spreadsheet_rows(export_variant: str = "default") -> List[Dict[str, Any]]:
    """Flat rows for Excel export (one row per call)."""
    count_crit = getattr(config, "METRICS_COUNT_CRITICAL_AS_VISHING_POSITIVE", True)
    rows: List[Dict[str, Any]] = []
    gt_pad = _blank_gt_cells() if export_variant == "gold_eval" else {}
    for uid in sorted(CALL_BUFFERS.keys()):
        buf = CALL_BUFFERS[uid]
        cl = buf.latest_classification
        exb = _explainability_export_blob(buf)
        tx = (buf.get_full_text() or "")[:31000]
        sub_resp = _subprompt_response_cells_for_xlsx(buf.latest_explainability)
        if (export_variant or "").strip().lower() == "gold_eval":
            ge_only = {
                "import_gt_0_1": _import_gt_numeric_for_export(buf),
                **_gold_eval_match_import_gt_cells(buf, sub_resp),
            }
        else:
            ge_only = {}
        base: Dict[str, Any] = {
            "call_uuid": uid,
            "ground_truth": _ground_truth_label(buf),
            "segment_count": len(buf.segments),
        }
        if not cl:
            rows.append(
                {
                    **base,
                    "risk_status": "PENDING",
                    "score": None,
                    "domain": "",
                    "category": "",
                    "label": "",
                    "pred_positive_fraud": "",
                    "bucket": "PENDING",
                    "classifier_analysis": "",
                    "transcript": tx,
                    "explainability_mode": (exb or {}).get("mode", "") if exb else "",
                    "explainability_steps": (exb or {}).get("steps_count", "") if exb else "",
                    "explainability_llm_only": (
                        "yes" if exb and exb.get("llm_only") else ("no" if exb else "")
                    ),
                    "explainability_custom_preview": (exb or {}).get("custom_prompt_preview", "")
                    if exb
                    else "",
                    **sub_resp,
                    **ge_only,
                    **gt_pad,
                    "explainability_llm": ((exb or {}).get("combined_llm_text") or "")[:31000] if exb else "",
                }
            )
            continue
        rs = cl.get("risk_status") or "ERROR"
        pred_pos = _prediction_positive_for_metrics(cl)
        if rs == "ERROR":
            bucket = "ERROR"
        elif rs == "VISHING":
            bucket = "VISHING"
        elif rs == "CRITICAL":
            bucket = "VISHING" if count_crit else "CRITICAL"
        elif rs == "SAFE":
            bucket = "SAFE"
        else:
            bucket = "ERROR"
        rows.append(
            {
                **base,
                "risk_status": rs,
                "score": cl.get("score"),
                "domain": cl.get("domain"),
                "category": cl.get("category"),
                "label": cl.get("label"),
                "pred_positive_fraud": pred_pos,
                "bucket": bucket,
                "classifier_analysis": (str(cl.get("analysis") or ""))[:31000],
                "transcript": tx,
                "explainability_mode": (exb or {}).get("mode", "") if exb else "",
                "explainability_steps": (exb or {}).get("steps_count", "") if exb else "",
                "explainability_llm_only": (
                    "yes" if exb and exb.get("llm_only") else ("no" if exb else "")
                ),
                "explainability_custom_preview": (exb or {}).get("custom_prompt_preview", "")
                if exb
                else "",
                **sub_resp,
                **ge_only,
                **gt_pad,
                "explainability_llm": ((exb or {}).get("combined_llm_text") or "")[:31000] if exb else "",
            }
        )
    return rows


_XLSX_EXPORT_COLUMNS_CORE: Tuple[str, ...] = (
    "call_uuid",
    "ground_truth",
    "segment_count",
    "risk_status",
    "score",
    "domain",
    "category",
    "label",
    "pred_positive_fraud",
    "bucket",
    "classifier_analysis",
    "transcript",
    "explainability_mode",
    "explainability_steps",
    "explainability_llm_only",
    "explainability_custom_preview",
)


def _xlsx_export_columns(export_variant: str) -> Tuple[str, ...]:
    ev = (export_variant or "default").strip().lower()
    core_x = _xlsx_export_columns_core_gold_eval() if ev == "gold_eval" else _XLSX_EXPORT_COLUMNS_CORE
    if ev == "gold_eval":
        mid = core_x + _xlsx_gold_rubric_mid_columns()
        return mid + _xlsx_gold_gt_column_names() + ("explainability_llm",)
    mid = core_x + _xlsx_custom_sub_column_names()
    return mid + ("explainability_llm",)


def _write_export_spreadsheet_xlsx(
    path: Path, rows: List[Dict[str, Any]], export_variant: str = "default"
) -> None:
    """Write one sheet using openpyxl only (no pandas)."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    cols = _xlsx_export_columns(export_variant)
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("openpyxl: no active worksheet")
    ws.title = "calls"
    for col_idx, name in enumerate(cols, 1):
        ws.cell(row=1, column=col_idx, value=_xlsx_header_display_name(name))
        letter = get_column_letter(col_idx)
        if name.startswith("subprompt_") and name.endswith("_match"):
            ws.column_dimensions[letter].width = 16
        elif name.startswith("subprompt_") or name.startswith("response_"):
            ws.column_dimensions[letter].width = 44
        elif name == "import_gt_0_1":
            ws.column_dimensions[letter].width = 14
        elif name.startswith("gt_"):
            ws.column_dimensions[letter].width = 22
        elif name == "annotator_notes":
            ws.column_dimensions[letter].width = 40
        elif name == "explainability_llm":
            ws.column_dimensions[letter].width = 56
        elif name == "transcript":
            ws.column_dimensions[letter].width = 48
    for row_idx, row in enumerate(rows, 2):
        for col_idx, name in enumerate(cols, 1):
            val: Any = _cell_value_for_xlsx(row.get(name, ""))
            ws.cell(row=row_idx, column=col_idx, value=val)
    ev_norm = (export_variant or "default").strip().lower()
    if ev_norm == "gold_eval":
        perf = _compute_gold_subprompt_match_performance(rows)
        # Summary sheet first so match % is visible on open.
        ws_rates = wb.create_sheet("subprompt_match_rates", 0)
        _fill_gold_match_rates_worksheet(ws_rates, perf, str(config.MODEL_NAME))
    wb.save(path)


def _ground_truth_label(buf: CallBuffer) -> Optional[str]:
    g = buf.ground_truth_is_vishing
    if g is True:
        return "vishing"
    if g is False:
        return "non_vishing"
    return None


def build_classification_results_export(export_variant: str = "default") -> Dict[str, Any]:
    """
    Snapshot of CALL_BUFFERS for JSON export (Save Results).
    Groups calls by risk_status; includes labeled metrics when GT is set.
    When METRICS_COUNT_CRITICAL_AS_VISHING_POSITIVE, CRITICAL rows use bucket VISHING (same fraud bin).
    """
    ev = (export_variant or "default").strip().lower()
    gt_extra_per_call = _gold_gt_json_fields(ev)
    count_crit = getattr(config, "METRICS_COUNT_CRITICAL_AS_VISHING_POSITIVE", True)
    by_risk: Dict[str, List[dict]] = {
        "VISHING": [],
        "CRITICAL": [],
        "SAFE": [],
        "ERROR": [],
        "PENDING": [],
    }
    uids_ordered = sorted(CALL_BUFFERS.keys())
    all_calls: List[dict] = []

    for uid in uids_ordered:
        buf = CALL_BUFFERS[uid]
        cl = buf.latest_classification
        if cl:
            _enrich_classification_dict(cl)
        _tx_export = (buf.get_full_text() or "")[:32000]
        _ex_blob = _explainability_export_blob(buf)
        gold_row_extra: Dict[str, Any] = {}
        if ev == "gold_eval":
            _sr = _subprompt_response_cells_for_xlsx(buf.latest_explainability)
            gold_row_extra = {
                "import_gt_0_1": _import_gt_numeric_json(buf),
                **_gold_eval_match_import_gt_cells(buf, _sr),
            }
        entry_base = {
            "call_uuid": uid,
            "ground_truth": _ground_truth_label(buf),
            "segment_count": len(buf.segments),
            "is_processing": buf.is_processing,
            "transcript": _tx_export,
            "explainability": _ex_blob,
            **gold_row_extra,
            **gt_extra_per_call,
        }
        if not cl:
            _pend_note = "no_classification_yet"
            if _ex_blob and int(_ex_blob.get("steps_count") or 0) > 0:
                _pend_note = "no_classification_yet; explainability_exported"
            entry = {**entry_base, "risk_status": "PENDING", "note": _pend_note}
            by_risk["PENDING"].append(entry)
            all_calls.append({**entry, "bucket": "PENDING"})
            continue
        rs = cl.get("risk_status") or "ERROR"
        pred_pos = _prediction_positive_for_metrics(cl)
        entry = {
            **entry_base,
            "risk_status": rs,
            "score": cl.get("score"),
            "domain": cl.get("domain"),
            "category": cl.get("category"),
            # `label` = YES if score ≥ FINAL threshold or (CRITICAL when counted as fraud-positive).
            "label": cl.get("label"),
            "pred_positive_fraud_metrics": pred_pos,
            "reviewed_full_conversation": cl.get("reviewed_full_conversation"),
            "heuristic_segment": cl.get("heuristic_segment"),
            "heuristic_full": cl.get("heuristic_full"),
            "segment_llm_steps": cl.get("segment_llm_steps"),
            "analysis": (str(cl.get("analysis") or ""))[:800],
            "classifier_analysis_full": (str(cl.get("analysis") or ""))[:32000],
        }
        bucket = "PENDING"
        if rs == "ERROR":
            bucket = "ERROR"
        elif rs == "VISHING":
            bucket = "VISHING"
        elif rs == "CRITICAL":
            bucket = "VISHING" if count_crit else "CRITICAL"
        elif rs == "SAFE":
            bucket = "SAFE"
        else:
            entry = {**entry, "note": f"unexpected_risk_status_{rs}"}
        by_risk[bucket].append(entry)
        all_calls.append({**entry, "bucket": bucket})

    metrics_uids = [u for u in uids_ordered if CALL_BUFFERS[u].ground_truth_is_vishing is not None]
    metrics = compute_confusion_for_call_uids(metrics_uids)

    summary = {
        "total_calls": len(CALL_BUFFERS),
        "vishing": len(by_risk["VISHING"]),
        "critical": len(by_risk["CRITICAL"]),
        "safe": len(by_risk["SAFE"]),
        "error": len(by_risk["ERROR"]),
        "pending": len(by_risk["PENDING"]),
        "critical_merged_into_vishing_bin": bool(count_crit),
    }

    out: Dict[str, Any] = {
        "export_format": "cyberaspis_classifier_run_v1",
        "saved_at_unix": time.time(),
        "saved_at_iso": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "app_title": app.title,
        "model": config.MODEL_NAME,
        "export_variant": ev,
        "config_snapshot": {
            "FINAL_VISHING_STOP_THRESHOLD": config.FINAL_VISHING_STOP_THRESHOLD,
            "SAFE_SCORE_MAX": config.SAFE_SCORE_MAX,
            "METRICS_COUNT_CRITICAL_AS_VISHING_POSITIVE": count_crit,
            "RECALL_HEURISTIC_FLOOR_THRESHOLD": config.RECALL_HEURISTIC_FLOOR_THRESHOLD,
            "RECALL_HEURISTIC_FLOOR_MAX_LLM": config.RECALL_HEURISTIC_FLOOR_MAX_LLM,
            "RECALL_HEURISTIC_FLOOR_VALUE": config.RECALL_HEURISTIC_FLOOR_VALUE,
            "BENIGN_SHORT_MAX_WORDS": getattr(config, "BENIGN_SHORT_MAX_WORDS", 48),
            "BENIGN_SHORT_MAX_HEURISTIC": getattr(config, "BENIGN_SHORT_MAX_HEURISTIC", 0.06),
            "BENIGN_SHORT_SCORE_CAP": getattr(config, "BENIGN_SHORT_SCORE_CAP", 0.26),
            "BENIGN_SHORT_PROMPT_NOTE": getattr(config, "BENIGN_SHORT_PROMPT_NOTE", True),
            "COMPLETION_ECHO_MAX_PARSED_PROB": getattr(
                config, "COMPLETION_ECHO_MAX_PARSED_PROB", 0.24
            ),
            "USE_CRITICAL_TRIAGE_SECOND_PASS": getattr(
                config, "USE_CRITICAL_TRIAGE_SECOND_PASS", True
            ),
            "CRITICAL_TRIAGE_SCORE_VISHING": getattr(
                config, "CRITICAL_TRIAGE_SCORE_VISHING", 0.58
            ),
            "CRITICAL_TRIAGE_SCORE_SAFE": getattr(config, "CRITICAL_TRIAGE_SCORE_SAFE", 0.14),
            "CRITICAL_TRIAGE_SCORE_UNCERTAIN": getattr(
                config, "CRITICAL_TRIAGE_SCORE_UNCERTAIN", 0.54
            ),
            "CLASSIFY_SAFE_LONG_RECHECK_ENABLED": getattr(
                config, "CLASSIFY_SAFE_LONG_RECHECK_ENABLED", True
            ),
            "CLASSIFY_SAFE_RECHECK_MIN_WORDS": getattr(
                config, "CLASSIFY_SAFE_RECHECK_MIN_WORDS", 220
            ),
            "CLASSIFY_SAFE_RECHECK_MAX_HEURISTIC": getattr(
                config, "CLASSIFY_SAFE_RECHECK_MAX_HEURISTIC", 0.12
            ),
            "CLASSIFY_SAFE_RECHECK_PROB_BELOW": getattr(
                config, "CLASSIFY_SAFE_RECHECK_PROB_BELOW", 0.22
            ),
            "BASELINE_BATCH_CONCURRENCY": getattr(config, "BASELINE_BATCH_CONCURRENCY", 100),
            "BASELINE_BATCH_CONCURRENCY_CAP": getattr(config, "BASELINE_BATCH_CONCURRENCY_CAP", 256),
            "USE_SAFE_DOMAIN_TRIAGE_RECHECK": getattr(
                config, "USE_SAFE_DOMAIN_TRIAGE_RECHECK", True
            ),
            "SAFE_DOMAIN_TRIAGE_MAX_PROB": getattr(config, "SAFE_DOMAIN_TRIAGE_MAX_PROB", 0.42),
            "SAFE_DOMAIN_TRIAGE_DOMAINS": sorted(
                getattr(config, "SAFE_DOMAIN_TRIAGE_DOMAINS", frozenset())
            ),
            "USE_SEGMENT_PROGRESSIVE_CLASSIFY": getattr(
                config, "USE_SEGMENT_PROGRESSIVE_CLASSIFY", False
            ),
            "VLLM_TRIAGE_COMPLETION_TOKENS": getattr(
                config, "VLLM_TRIAGE_COMPLETION_TOKENS", 48
            ),
            "MAIN_PROMPT_MAX_NEW_TOKENS": getattr(config, "MAIN_PROMPT_MAX_NEW_TOKENS", 512),
            "SUBPROMPT_MAX_NEW_TOKENS": getattr(config, "SUBPROMPT_MAX_NEW_TOKENS", 24),
            "MAX_INPUT_TOKENS_PER_CALL": getattr(config, "MAX_INPUT_TOKENS_PER_CALL", 5500),
            "MAX_OUTPUT_TOKENS_PER_CALL": getattr(config, "MAX_OUTPUT_TOKENS_PER_CALL", 24),
            "MAX_OUTPUT_TOKENS_MAIN_PROMPT": getattr(config, "MAX_OUTPUT_TOKENS_MAIN_PROMPT", 512),
            "CUSTOM_EXPLAIN_SUB_STYLE": getattr(config, "CUSTOM_EXPLAIN_SUB_STYLE", "yesno"),
        },
        "summary": summary,
        "metrics_labeled_only": metrics,
        "by_risk": by_risk,
        "all_calls": all_calls,
    }
    if ev == "gold_eval":
        out["gold_rubric"] = {
            "id": gold_subprompt_rubric.RUBRIC_ID,
            "subprompts": [
                {"key": k, "text": t}
                for k, t in zip(
                    gold_subprompt_rubric.SUBPROMPT_KEYS,
                    gold_subprompt_rubric.DEFAULT_SUBPROMPTS,
                    strict=True,
                )
            ],
            "import_gt_0_1_note": "1 = import label vishing, 0 = not, null = no import label.",
            "subprompt_match_note": (
                "subprompt_N_match: YES if LLM YES/NO aligns with import_gt_0_1 (1→expect YES, 0→expect NO); "
                "NO if not; UNPARSED if LLM output unclear; empty if no import label. "
                "Naive for subprompts 2–6 (indicator vs global label)."
            ),
            "note_el": "gt_1..gt_N: χειροκίνητα. subprompt_N_match: αυτόματο YES/NO έναντι import GT.",
        }
        sheet_rows = _build_export_spreadsheet_rows("gold_eval")
        out["gold_subprompt_match_performance"] = _compute_gold_subprompt_match_performance(
            sheet_rows
        )
    return out


@app.post("/save_classification_results")
async def save_classification_results(request: Request):
    """
    Write current dashboard classifications to Project_v2/classifier_run_logs/*.json
    for offline review and fine-tuning. When CRITICAL counts as fraud, export bucket is VISHING.

    Optional JSON body: {"export_variant": "gold_eval"} — includes GT columns in XLSX/JSON (after gold-eval run).
    """
    export_variant = "default"
    try:
        body = await request.json()
        if isinstance(body, dict) and body.get("export_variant") == "gold_eval":
            export_variant = "gold_eval"
    except Exception:
        pass
    try:
        RESULTS_EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        return {"ok": False, "error": f"Cannot create export dir: {e}"}
    if not CALL_BUFFERS:
        return {
            "ok": False,
            "error": "No calls loaded — import conversations first (classifier and/or explainability runs are optional).",
        }
    payload = build_classification_results_export(export_variant)
    ts = time.strftime("%Y%m%dT%H%M%SZ", time.gmtime())
    suffix = "_gold_eval" if export_variant == "gold_eval" else ""
    fname = f"classifier_run_{ts}{suffix}.json"
    path = RESULTS_EXPORT_DIR / fname
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except OSError as e:
        logger.exception("Save classification results failed")
        return {"ok": False, "error": str(e)}
    rel = f"classifier_run_logs/{fname}"
    logger.info("Saved classification export to %s", path)

    xlsx_name = f"classifier_run_{ts}{suffix}.xlsx"
    xlsx_path = RESULTS_EXPORT_DIR / xlsx_name
    try:
        _write_export_spreadsheet_xlsx(
            xlsx_path, _build_export_spreadsheet_rows(export_variant), export_variant
        )
        logger.info("Saved spreadsheet export to %s", xlsx_path)
    except Exception as e:
        logger.exception("Save XLSX failed: %s", e)
        return {
            "ok": False,
            "error": f"JSON saved but XLSX failed: {e}",
            "filename": fname,
            "absolute_path": str(path.resolve()),
        }

    return {
        "ok": True,
        "filename": fname,
        "xlsx_filename": xlsx_name,
        "relative_path": rel,
        "xlsx_relative_path": f"classifier_run_logs/{xlsx_name}",
        "absolute_path": str(path.resolve()),
        "xlsx_absolute_path": str(xlsx_path.resolve()),
        "summary": payload["summary"],
        "metrics_labeled_only": payload["metrics_labeled_only"],
        "export_variant": export_variant,
    }


@app.get("/download_export/{filename}")
async def download_export(filename: str):
    """Download a saved classifier_run_* .json or .xlsx from classifier_run_logs/."""
    if not _EXPORT_FILENAME_SAFE.match(filename):
        raise HTTPException(status_code=400, detail="Invalid export filename")
    path = (RESULTS_EXPORT_DIR / filename).resolve()
    try:
        root = RESULTS_EXPORT_DIR.resolve()
    except OSError:
        raise HTTPException(status_code=500, detail="Export dir unavailable")
    if path.parent != root or not path.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    media = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if filename.endswith(".xlsx")
        else "application/json"
    )
    return FileResponse(path, filename=filename, media_type=media)


@app.get("/health", response_model=HealthResponse)
async def health_check():
    vllm_ok = await vllm_client.health_check() if vllm_client else False
    agent_summary = (
        f"Mistral-7B Instruct (local vLLM) — {config.MODEL_NAME}"
    )
    return HealthResponse(
      status="ok" if vllm_ok else "degraded",
      vllm_healthy=vllm_ok,
      model=config.MODEL_NAME,
      env=config.ENV,
      inference_engine="vLLM",
      model_source="Hugging Face Hub",
      agent_summary=agent_summary,
      vllm_base_url=config.VLLM_BASE_URL,
    )


@app.post("/run_baseline_sample")
async def run_baseline_sample(request: Request):
    """
    Run baseline4 behavioral fraud classifier on up to N calls (default 10).
    Updates each call's classification with score and STATE (SAFE/CRITICAL/VISHING).
    """
    global vllm_client
    if not vllm_client:
        return {"ok": False, "error": "vLLM client not ready"}
    try:
        body = await request.json()
    except Exception:
        body = {}
    limit = int(body.get("limit", 10))
    limit = max(1, min(limit, 500))

    # Buffers that have text
    candidates = [
        (uid, buf)
        for uid, buf in list(CALL_BUFFERS.items())
        if buf.get_full_text().strip()
    ]
    to_run = candidates[:limit]
    if not to_run:
        await broadcast_call_list()
        return {
            "ok": True,
            "classified": 0,
            "message": "No calls with text",
            "candidates_count": 0,
            "metrics": compute_confusion_for_call_uids([]),
        }

    classified = 0
    first_error = None
    cap = int(getattr(config, "BASELINE_BATCH_CONCURRENCY_CAP", 256))
    conc = int(getattr(config, "BASELINE_BATCH_CONCURRENCY", 100))
    conc = max(1, min(conc, cap))
    for batch_start in range(0, len(to_run), conc):
        batch = to_run[batch_start : batch_start + conc]

        async def _one(pair: Tuple[str, CallBuffer]) -> Optional[str]:
            u, b = pair
            return await _baseline_classify_single_buffer(u, b, progress_queue=None)

        errors = await asyncio.gather(*[_one(pair) for pair in batch])
        for err in errors:
            classified += 1
            if err and first_error is None:
                first_error = err

    await broadcast_call_list()
    uids_done = [uid for uid, _ in to_run]
    metrics = compute_confusion_for_call_uids(uids_done)
    out = {
        "ok": True,
        "classified": classified,
        "candidates_count": len(to_run),
        "metrics": metrics,
    }
    if first_error:
        out["error"] = first_error
    return out


@app.get("/run_baseline_stream")
async def run_baseline_stream(limit: int = 100):
    """
    Run classifier on up to `limit` calls and stream progress as SSE.
    Events: {started, total}, {done, total}, {finished, total, metrics}.
    """
    global vllm_client
    limit = max(1, min(limit, 500))

    async def event_stream():
        if not vllm_client:
            logger.error("run_baseline_stream: vLLM client not ready")
            yield f"data: {json.dumps({'done': 0, 'total': 0, 'finished': True, 'error': 'vLLM not ready'})}\n\n"
            return
        try:
            ok = await vllm_client.health_check()
            if not ok:
                # Don't block: run classification anyway; first request will fail with clear error if vLLM is down
                logger.warning("vLLM health check failed; will attempt classification (first call may fail)")
        except Exception as e:
            logger.warning("vLLM health check error: %s; will attempt classification", e)
        candidates = [
            (uid, buf)
            for uid, buf in list(CALL_BUFFERS.items())
            if buf.get_full_text().strip()
        ]
        to_run = candidates[:limit]
        total = len(to_run)
        if total == 0:
            yield f"data: {json.dumps({'done': 0, 'total': 0, 'finished': True, 'message': 'No calls with text — import conversations first', 'metrics': compute_confusion_for_call_uids([])})}\n\n"
            return
        yield f"data: {json.dumps({'started': True, 'total': total})}\n\n"
        cap = int(getattr(config, "BASELINE_BATCH_CONCURRENCY_CAP", 256))
        conc = max(1, min(int(getattr(config, "BASELINE_BATCH_CONCURRENCY", 100)), cap))
        done_n = 0
        i = 0
        while i < total:
            batch = to_run[i : i + conc]
            for j, (uid, buffer) in enumerate(batch):
                yield f"data: {json.dumps({'call_started': True, 'call_index': i + j + 1, 'total_calls': total, 'call_uuid': uid, 'batch_concurrency': conc})}\n\n"

            if conc == 1:
                uid, buffer = batch[0]
                call_index = i + 1
                if buffer.is_vishing_final:
                    logger.info(f"Baseline4 stream skipped {uid}: already finalized as vishing")
                else:
                    prog_q: asyncio.Queue = asyncio.Queue()
                    cls_task = asyncio.create_task(
                        _baseline_classify_single_buffer(uid, buffer, progress_queue=prog_q)
                    )
                    while not cls_task.done():
                        try:
                            tick = await asyncio.wait_for(prog_q.get(), timeout=0.35)
                            yield f"data: {json.dumps({'segment_tick': True, 'call_index': call_index, 'total_calls': total, **tick})}\n\n"
                        except asyncio.TimeoutError:
                            pass
                    while True:
                        try:
                            tick = prog_q.get_nowait()
                            yield f"data: {json.dumps({'segment_tick': True, 'call_index': call_index, 'total_calls': total, **tick})}\n\n"
                        except asyncio.QueueEmpty:
                            break
                    await cls_task
                if i == 0:
                    cl0 = buffer.latest_classification or {}
                    logger.info(
                        f"Baseline4 stream first call {uid}: progressive={cl0.get('used_segment_progressive')} "
                        f"llm_steps={cl0.get('segment_llm_steps')}"
                    )
                done_n += 1
                yield f"data: {json.dumps({'done': done_n, 'total': total})}\n\n"
                await broadcast_call_list()
            else:
                await asyncio.gather(
                    *[
                        _baseline_classify_single_buffer(u, b, progress_queue=None)
                        for u, b in batch
                    ]
                )
                for _ in batch:
                    done_n += 1
                    yield f"data: {json.dumps({'done': done_n, 'total': total})}\n\n"
                    await broadcast_call_list()
            i += len(batch)
        uids_ordered = [uid for uid, _ in to_run]
        final_metrics = compute_confusion_for_call_uids(uids_ordered)
        yield f"data: {json.dumps({'done': total, 'total': total, 'finished': True, 'metrics': final_metrics})}\n\n"

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


async def _explainability_event_stream(
    limit: int,
    custom_prompt: Optional[str] = None,
    custom_subprompts: Optional[List[str]] = None,
    custom_prompt_preserve_output: bool = False,
    gold_eval_run: bool = False,
):
    """
    SSE: Built-in Prompt #4 explainability (with classifier), OR custom modes that only
    stream **Mistral** completions: one full prompt per call, or N chained sub-prompts per call.
    """
    global vllm_client
    limit = max(1, min(limit, 500))
    custom_str = (custom_prompt or "").strip()
    norm_subs = custom_subprompts
    use_custom_subs = bool(norm_subs)
    use_custom_full = bool(custom_str) and not use_custom_subs
    preserve_custom_output = bool(
        custom_prompt_preserve_output and use_custom_full and len(custom_str) > 0
    )
    n_steps_full = baseline4.explainability_step_count()
    if use_custom_full:
        n_steps_stream = 1
    elif use_custom_subs and norm_subs:
        n_steps_stream = len(norm_subs)
    else:
        n_steps_stream = n_steps_full

    if not vllm_client:
        logger.error("run_baseline_explainability_stream: vLLM client not ready")
        yield f"data: {json.dumps({'done': 0, 'total': 0, 'finished': True, 'error': 'vLLM not ready'})}\n\n"
        return
    try:
        ok = await vllm_client.health_check()
        if not ok:
            logger.warning("vLLM health check failed; explainability stream will attempt calls anyway")
    except Exception as e:
        logger.warning("vLLM health check error: %s; proceeding", e)

    candidates = [
        (uid, buf)
        for uid, buf in list(CALL_BUFFERS.items())
        if buf.get_full_text().strip()
    ]
    to_run = candidates[:limit]
    total = len(to_run)
    if total == 0:
        yield f"data: {json.dumps({'done': 0, 'total': 0, 'finished': True, 'message': 'No calls with text — import conversations first', 'metrics': compute_confusion_for_call_uids([])})}\n\n"
        return

    cap = int(getattr(config, "BASELINE_BATCH_CONCURRENCY_CAP", 256))
    conc = max(1, min(int(getattr(config, "BASELINE_BATCH_CONCURRENCY", 100)), cap))
    if not use_custom_full and not use_custom_subs:
        conc = 1

    yield f"data: {json.dumps({'started': True, 'total': total, 'explainability': True, 'steps_per_call': n_steps_stream, 'explainability_custom': use_custom_full, 'explainability_custom_subprompts': use_custom_subs, 'explainability_custom_full_llm_only': use_custom_full, 'explainability_mistral_explain_only': use_custom_full or use_custom_subs, 'batch_concurrency': conc, 'gold_eval_run': bool(gold_eval_run)})}\n\n"

    stream_explain_variant = (
        "custom_full" if use_custom_full else ("custom_subprompts" if use_custom_subs else "baseline")
    )

    done_n = 0
    for batch_start in range(0, total, conc):
        batch = to_run[batch_start : batch_start + conc]

        for j, (uid, buffer) in enumerate(batch):
            call_index = batch_start + j + 1
            yield f"data: {json.dumps({'call_started': True, 'call_index': call_index, 'total_calls': total, 'call_uuid': uid, 'mode': 'explainability', 'batch_concurrency': conc, 'explain_variant': stream_explain_variant})}\n\n"

        if use_custom_full:
            title_el = (
                "Explainability — μοτίβα απάτης (JSON, 1× LLM)"
                if preserve_custom_output
                else "Προσαρμοσμένο prompt (ενιαίο βήμα LLM)"
            )
            mt = baseline4.custom_explain_main_max_tokens()
            preview = custom_str[:500] + ("…" if len(custom_str) > 500 else "")
            _out_fmt = "json" if preserve_custom_output else "prose"

            async def _custom_llm(
                item: Tuple[str, CallBuffer, int],
            ) -> Tuple[str, CallBuffer, int, Optional[str], bool]:
                uid, buffer, call_index = item
                full_text = baseline4.segments_full_transcript(buffer.segments).strip()
                if not full_text:
                    return uid, buffer, call_index, None, True
                try:
                    prompt = baseline4.build_custom_explainability_prompt(
                        custom_str, full_text, output_format=_out_fmt
                    )
                    result = await _vllm_complete_with_fallback(prompt, max_tokens=mt)
                    raw = _completion_text_from_result(result)
                    if not preserve_custom_output:
                        raw = baseline4.sanitize_explainability_completion(raw, full_text)
                        raw = baseline4.sanitize_custom_full_completion(raw)
                    else:
                        raw = (raw or "").strip()
                except Exception as e:
                    logger.error("Custom explainability for %s: %s", uid, e, exc_info=True)
                    raw = f"[σφάλμα vLLM] {e}"
                return uid, buffer, call_index, raw, False

            llm_inputs = [(uid, buf, batch_start + j + 1) for j, (uid, buf) in enumerate(batch)]
            llm_results = await asyncio.gather(*[_custom_llm(x) for x in llm_inputs])
            llm_results = sorted(llm_results, key=lambda x: x[2])

            for uid, buffer, call_index, raw, is_empty in llm_results:
                if is_empty:
                    buffer.latest_explainability = {
                        "steps": [],
                        "updated_at": time.time(),
                        "note": "empty_transcript",
                        "mode": "custom",
                        "llm_only": True,
                    }
                else:
                    rec = {
                        "step_index": 0,
                        "step_key": "custom",
                        "step_title": title_el,
                        "llm_question": custom_str,
                        "llm_raw": raw,
                    }
                    buffer.latest_explainability = {
                        "steps": [rec],
                        "updated_at": time.time(),
                        "mode": "custom",
                        "custom_prompt_preview": preview,
                        "llm_only": True,
                    }
                    yield f"data: {json.dumps({'explain_step': True, 'call_uuid': uid, 'call_index': call_index, 'total_calls': total, 'step_index': 0, 'step_total': 1, 'step_key': 'custom', 'step_title': title_el, 'llm_question': custom_str, 'llm_raw': raw})}\n\n"

            await broadcast_call_list()
            done_n += len(batch)
            yield f"data: {json.dumps({'done': done_n, 'total': total})}\n\n"
            await broadcast_call_list()

            if batch_start == 0 and to_run:
                u0 = to_run[0][0]
                b0 = CALL_BUFFERS.get(u0)
                if b0 and b0.latest_explainability:
                    logger.info(
                        "Custom full-prompt (LLM-only) first call %s: explain steps=%s",
                        u0,
                        len((b0.latest_explainability or {}).get("steps") or []),
                    )

        elif use_custom_subs:
            assert norm_subs is not None
            n_sub = len(norm_subs)
            subs_preview = [s[:120] + ("…" if len(s) > 120 else "") for s in norm_subs]
            llm_inputs = [(uid, buf, batch_start + j + 1) for j, (uid, buf) in enumerate(batch)]

            valid_items: List[Tuple[str, CallBuffer, int]] = []

            for uid, buffer, call_index in llm_inputs:
                ft = baseline4.segments_full_transcript(buffer.segments).strip()
                if not ft:
                    buffer.latest_explainability = {
                        "steps": [],
                        "updated_at": time.time(),
                        "note": "empty_transcript",
                        "mode": "custom_subprompts",
                        "custom_subprompts_preview": subs_preview,
                        "llm_only": True,
                        **_gold_eval_explain_fields(gold_eval_run),
                    }
                else:
                    valid_items.append((uid, buffer, call_index))

            acc_uids: Dict[str, Dict[str, Any]] = {uid: {"prior": [], "steps": []} for uid, _, _ in valid_items}

            mt_sub = baseline4.custom_explain_sub_max_tokens()
            for step_idx in range(n_sub):
                focus = norm_subs[step_idx]
                step_key = f"custom_sub_{step_idx}"
                if gold_eval_run:
                    title_el = f"Ground Truth ({step_idx + 1}/{n_sub})"
                else:
                    title_el = f"Προσαρμοσμένο υπο-prompt ({step_idx + 1}/{n_sub})"

                async def _subs_step_one(
                    item: Tuple[str, CallBuffer, int],
                    _step_idx: int = step_idx,
                    _focus: str = focus,
                ) -> Tuple[str, CallBuffer, int, str]:
                    uid, buffer, call_index = item
                    ft = baseline4.segments_full_transcript(buffer.segments).strip()
                    prior = list(acc_uids[uid]["prior"])
                    try:
                        prompt = baseline4.build_custom_explain_chain_turn(
                            ft,
                            _focus,
                            prior[:_step_idx],
                            turn_index=_step_idx,
                        )
                        result = await _vllm_complete_with_fallback(prompt, max_tokens=mt_sub)
                        raw = _completion_text_from_result(result)
                    except Exception as e:
                        logger.error(
                            "Custom explain chain step %s for %s: %s",
                            _step_idx,
                            uid,
                            e,
                            exc_info=True,
                        )
                        raw = f"[σφάλμα vLLM] {e}"
                    else:
                        raw = baseline4.sanitize_explainability_completion(raw, ft)
                        if baseline4.custom_explain_sub_style_is_yesno():
                            raw = baseline4.normalize_custom_sub_yesno_completion(raw)
                    return uid, buffer, call_index, raw

                step_results = await asyncio.gather(*[_subs_step_one(x) for x in valid_items])
                step_results.sort(key=lambda x: x[2])

                for uid, buffer, call_index, raw in step_results:
                    acc_uids[uid]["prior"].append(raw)
                    rec = {
                        "step_index": step_idx,
                        "step_key": step_key,
                        "step_title": title_el,
                        "llm_question": focus,
                        "llm_raw": raw,
                    }
                    acc_uids[uid]["steps"].append(rec)
                    buffer.latest_explainability = {
                        "steps": list(acc_uids[uid]["steps"]),
                        "updated_at": time.time(),
                        "mode": "custom_subprompts",
                        "custom_subprompts_preview": subs_preview,
                        "llm_only": True,
                        **_gold_eval_explain_fields(gold_eval_run),
                    }
                    yield f"data: {json.dumps({'explain_step': True, 'call_uuid': uid, 'call_index': call_index, 'total_calls': total, 'step_index': step_idx, 'step_total': n_sub, 'step_key': step_key, 'step_title': title_el, 'llm_question': focus, 'llm_raw': raw})}\n\n"

                await broadcast_call_list()

            await broadcast_call_list()
            done_n += len(batch)
            yield f"data: {json.dumps({'done': done_n, 'total': total})}\n\n"
            await broadcast_call_list()

            if batch_start == 0 and to_run:
                u0 = to_run[0][0]
                b0 = CALL_BUFFERS.get(u0)
                if b0 and b0.latest_explainability:
                    logger.info(
                        "Custom sub-prompts (Mistral-only) first call %s: steps=%s",
                        u0,
                        len((b0.latest_explainability or {}).get("steps") or []),
                    )

        else:
            uid, buffer = batch[0]
            call_index = batch_start + 1
            steps_out: List[dict] = []
            prior_raw: List[str] = []
            full_text = baseline4.segments_full_transcript(buffer.segments).strip()

            if not full_text:
                buffer.latest_explainability = {
                    "steps": [],
                    "updated_at": time.time(),
                    "note": "empty_transcript",
                    "mode": "baseline_steps",
                }
                await broadcast_call_list()
                if not buffer.is_vishing_final:
                    prog_q: asyncio.Queue = asyncio.Queue()
                    cls_task = asyncio.create_task(
                        _baseline_classify_single_buffer(uid, buffer, progress_queue=prog_q)
                    )
                    while not cls_task.done():
                        try:
                            tick = await asyncio.wait_for(prog_q.get(), timeout=0.35)
                            yield f"data: {json.dumps({'segment_tick': True, 'call_index': call_index, 'total_calls': total, **tick})}\n\n"
                        except asyncio.TimeoutError:
                            pass
                    while True:
                        try:
                            tick = prog_q.get_nowait()
                            yield f"data: {json.dumps({'segment_tick': True, 'call_index': call_index, 'total_calls': total, **tick})}\n\n"
                        except asyncio.QueueEmpty:
                            break
                    await cls_task
            else:
                for step_idx in range(max(0, n_steps_full - 1)):
                    spec = baseline4.explainability_step_spec(step_idx)
                    mt = baseline4.explainability_max_tokens_for_step(step_idx)
                    try:
                        prompt = baseline4.build_explainability_prompt(
                            full_text, step_idx, prior_raw, use_full_text=True
                        )
                        result = await _vllm_complete_with_fallback(prompt, max_tokens=mt)
                        raw = _completion_text_from_result(result)
                    except Exception as e:
                        logger.error("Explainability step %s for %s: %s", step_idx, uid, e, exc_info=True)
                        raw = f"[σφάλμα vLLM] {e}"
                    else:
                        raw = baseline4.sanitize_explainability_completion(raw, full_text)

                    _llm_q = _explain_step_llm_question_baseline(spec)
                    rec = {
                        "step_index": step_idx,
                        "step_key": spec["key"],
                        "step_title": spec["title_el"],
                        "llm_question": _llm_q,
                        "llm_raw": raw,
                    }
                    steps_out.append(rec)
                    prior_raw.append(raw)
                    yield f"data: {json.dumps({'explain_step': True, 'call_uuid': uid, 'call_index': call_index, 'total_calls': total, 'step_index': step_idx, 'step_total': n_steps_full, 'step_key': spec['key'], 'step_title': spec['title_el'], 'llm_question': _llm_q, 'llm_raw': raw})}\n\n"

                if buffer.is_vishing_final:
                    logger.info(f"Explainability stream: classifier skipped {uid} (already finalized vishing)")
                else:
                    prog_q = asyncio.Queue()
                    cls_task = asyncio.create_task(
                        _baseline_classify_single_buffer(uid, buffer, progress_queue=prog_q)
                    )
                    while not cls_task.done():
                        try:
                            tick = await asyncio.wait_for(prog_q.get(), timeout=0.35)
                            yield f"data: {json.dumps({'segment_tick': True, 'call_index': call_index, 'total_calls': total, **tick})}\n\n"
                        except asyncio.TimeoutError:
                            pass
                    while True:
                        try:
                            tick = prog_q.get_nowait()
                            yield f"data: {json.dumps({'segment_tick': True, 'call_index': call_index, 'total_calls': total, **tick})}\n\n"
                        except asyncio.QueueEmpty:
                            break
                    await cls_task

                if n_steps_full >= 1:
                    last_idx = n_steps_full - 1
                    spec = baseline4.explainability_step_spec(last_idx)
                    mt = baseline4.explainability_max_tokens_for_step(last_idx)
                    cl = buffer.latest_classification or {}
                    classifier_hint = {
                        "score": cl.get("score"),
                        "risk_status": cl.get("risk_status"),
                        "label": cl.get("label"),
                        "domain": cl.get("domain"),
                    }
                    if not any(v is not None for v in classifier_hint.values()):
                        classifier_hint = None
                    try:
                        prompt = baseline4.build_explainability_prompt(
                            full_text,
                            last_idx,
                            prior_raw,
                            use_full_text=True,
                            classifier_hint=classifier_hint,
                        )
                        result = await _vllm_complete_with_fallback(prompt, max_tokens=mt)
                        raw = _completion_text_from_result(result)
                    except Exception as e:
                        logger.error("Explainability final step for %s: %s", uid, e, exc_info=True)
                        raw = f"[σφάλμα vLLM] {e}"
                    else:
                        raw = baseline4.sanitize_explainability_completion(raw, full_text)

                    _llm_q_f = _explain_step_llm_question_baseline(spec)
                    rec = {
                        "step_index": last_idx,
                        "step_key": spec["key"],
                        "step_title": spec["title_el"],
                        "llm_question": _llm_q_f,
                        "llm_raw": raw,
                    }
                    steps_out.append(rec)
                    yield f"data: {json.dumps({'explain_step': True, 'call_uuid': uid, 'call_index': call_index, 'total_calls': total, 'step_index': last_idx, 'step_total': n_steps_full, 'step_key': spec['key'], 'step_title': spec['title_el'], 'llm_question': _llm_q_f, 'llm_raw': raw})}\n\n"

                buffer.latest_explainability = {
                    "steps": steps_out,
                    "updated_at": time.time(),
                    "mode": "baseline_steps",
                }
            await broadcast_call_list()
            done_n += 1
            yield f"data: {json.dumps({'done': done_n, 'total': total})}\n\n"
            await broadcast_call_list()

            if batch_start == 0:
                u0, b0 = to_run[0]
                b0 = CALL_BUFFERS.get(u0, b0)
                if b0 and b0.latest_classification:
                    cl0 = b0.latest_classification or {}
                    logger.info(
                        "Explainability+classifier first call %s: progressive=%s llm_steps=%s",
                        u0,
                        cl0.get("used_segment_progressive"),
                        cl0.get("segment_llm_steps"),
                    )

    uids_ordered = [uid for uid, _ in to_run]
    final_metrics = compute_confusion_for_call_uids(uids_ordered)
    yield f"data: {json.dumps({'done': total, 'total': total, 'finished': True, 'metrics': final_metrics})}\n\n"


@app.get("/run_baseline_explainability_stream")
async def run_baseline_explainability_stream(limit: int = 100):
    """
    For each call: built-in Prompt #4 (sub-questions + classifier + synthesis), unless POST uses
    custom_prompt or custom_subprompts (Mistral explainability only, no classifier).

    Custom modes use batched concurrency; built-in flow uses concurrency 1.
    """
    return StreamingResponse(
        _explainability_event_stream(limit),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/run_baseline_explainability_stream")
async def run_baseline_explainability_stream_post(body: ExplainabilityStreamBody):
    """
    Same SSE as GET. JSON body may include:
    - custom_prompt: one instruction → single Mistral completion per call (+ transcript).
    - custom_subprompts: 1..N non-empty strings → N chained Mistral completions per call (no classifier).
    If both are sent, custom_subprompts wins when valid.
    - gold_eval_run: if true, ignores custom fields and runs fixed rubric sub-prompts (gold-eval export).
    """
    gold_eval_run = bool(body.gold_eval_run)
    if gold_eval_run:
        subs = _normalize_explainability_custom_subprompts(
            list(gold_subprompt_rubric.DEFAULT_SUBPROMPTS)
        )
        cust_prompt: Optional[str] = None
        preserve_out = False
    else:
        subs = _normalize_explainability_custom_subprompts(body.custom_subprompts)
        cust_prompt = body.custom_prompt
        preserve_out = body.custom_prompt_preserve_output
    return StreamingResponse(
        _explainability_event_stream(
            body.limit,
            custom_prompt=cust_prompt,
            custom_subprompts=subs,
            custom_prompt_preserve_output=preserve_out,
            gold_eval_run=gold_eval_run,
        ),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# --- Startup / Shutdown ---
@app.on_event("startup")
async def startup_event():
    global vllm_client, action_manager

    loop = asyncio.get_event_loop()
    loop.set_default_executor(ThreadPoolExecutor(max_workers=100))

    vllm_client = VLLMClient()
    action_manager = ActionManager()

    logger.info(
        f"Classifier Service started (LLM-only) | env={config.ENV} | "
        f"vllm={config.VLLM_BASE_URL} | model={config.MODEL_NAME} | "
        f"actions={action_manager.base_url} enabled={actions_enabled}"
    )
    asyncio.create_task(cleanup_task())


@app.on_event("shutdown")
async def shutdown_event():
    global vllm_client, action_manager
    if vllm_client:
        await vllm_client.close()
    if action_manager:
        await action_manager.close()


async def cleanup_task():
    """Remove stale call buffers every 5 minutes."""
    logger.info("Cleanup task started")
    while True:
        try:
            await asyncio.sleep(300)
            now = time.time()
            to_delete = [
                uid
                for uid, buffer in CALL_BUFFERS.items()
                if now - buffer.last_update > 1800
            ]
            if to_delete:
                for uid in to_delete:
                    del CALL_BUFFERS[uid]
                    if action_manager:
                        action_manager.cleanup(uid)
                logger.info(
                    f"Cleaned up {len(to_delete)} old call buffers. "
                    f"Remaining: {len(CALL_BUFFERS)}"
                )
        except Exception as e:
            logger.error(f"Error in cleanup task: {e}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=config.CLASSIFIER_PORT)
